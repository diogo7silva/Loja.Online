import openpyxl
from user import User


class linhaCarrinho:
    def __init__(self):
        self.reset()

    def reset(self):
        self.folha = ''
        self.linha = 0
        self.codigo = ''
        self.artigo = ''
        self.quantidade = 1
        self.custo = 0

    def total(self):
        return self.quantidade * self.custo


class lojaOnline:
    livro = openpyxl.load_workbook('db/Produtos.xlsx')
    produtos = livro.worksheets
    items_linha = 3
    items_pagina = 9

    def __init__(self):
        self.reset()

    def reset(self):
        self.usr = User()
        self.encomendar = False
        self.carrinho = []
        self.unidades = {}
        self.fatura = ''
        self.pagina = 0

    def atualiza(self):
        for li in range(self.carrinho.__len__()):
            aux = self.unidades[(self.carrinho[li].folha, self.carrinho[li].linha)]
            self.carrinho[li].quantidade = aux

    def addcarrinho(self, produto, linha):
        if (produto, linha) in self.unidades:
            self.unidades[produto, linha] += 1
            self.atualiza()
        else:
            tmp = linhaCarrinho()
            tmp.folha = produto
            tmp.linha = linha
            folha = self.livro[tmp.folha]
            tmp.codigo = folha.cell(row=tmp.linha, column=1).value
            for coluna in range(2, folha.max_column):
                tmp.artigo += str(folha.cell(row=1, column=coluna).value) + ': '
                tmp.artigo += str(folha.cell(row=tmp.linha, column=coluna).value) + '; '
            tmp.custo = folha.cell(row=tmp.linha, column=folha.max_column).value
            self.carrinho.append(tmp)
            self.unidades[produto, linha] = 1

    def subcarrinho(self, linha):
        chave = (self.carrinho[linha].folha, self.carrinho[linha].linha)
        self.unidades[chave] -= 1
        self.atualiza()
        if self.unidades[chave] == 0:
            del self.carrinho[linha]
            del self.unidades[chave]
        if self.carrinho.__len__() <= self.pagina * self.items_pagina:
            self.pagina -= 1

    def encomenda(self):
        from datetime import datetime
        data = datetime.now()
        db = 'db/Encomendas.xlsx'
        try:
            livro = openpyxl.load_workbook(db)
        except:
            livro = openpyxl.Workbook()
        folha = livro.active
        for li in range(self.carrinho.__len__()):
            aux = self.carrinho[li]
            dados = [self.fatura[0:-4], data, self.usr.id]
            dados += [aux.folha, aux.linha, aux.codigo, aux.artigo]
            folha.append(dados + [aux.quantidade, aux.custo, aux.total()])
        livro.save(db)

    def total(self):
        valor = 0
        for linha in range(self.carrinho.__len__()):
            valor += self.carrinho[linha].total()
        return valor

    @staticmethod
    def moeda(valor):
        import locale
        locale.setlocale(locale.LC_ALL, '')
        return locale.currency(valor, grouping=True)
