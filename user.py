import openpyxl


class User:
    db = 'db/Utilizadores.xlsx'

    def __init__(self):
        self.reset()

    def reset(self):
        self.id = None
        self.cliente = ''
        self.email = ''
        self.password = ''
        self.nif = ''
        self.nome = ''
        self.morada = ''
        self.init_db()

    def init_db(self):
        import os
        try:
            os.stat(self.db)
        except:
            wb = openpyxl.Workbook()
            wb.save(self.db)

    def set(self, usr):
        livro = openpyxl.load_workbook(self.db)
        folha = livro.active
        for li in range(1, folha.max_row + 1):
            if folha.cell(row=li, column=1).value == usr:
                self.id = usr
                self.cliente = 'C' + str(li).zfill(6)
                self.email = folha.cell(row=li, column=2).value
                self.password = folha.cell(row=li, column=3).value
                if folha.cell(row=li, column=4).value:
                    self.nif = folha.cell(row=li, column=4).value
                if folha.cell(row=li, column=5).value:
                    self.nome = folha.cell(row=li, column=5).value
                if folha.cell(row=li, column=6).value:
                    self.morada = folha.cell(row=li, column=6).value

    def save(self, id, email, password):
        livro = openpyxl.load_workbook(self.db)
        folha = livro.active
        folha.append([id, email, self.code(password)])
        livro.save(self.db)

    def update(self):
        livro = openpyxl.load_workbook(self.db)
        folha = livro.active
        for li in range(1, folha.max_row + 1):
            if folha.cell(row=li, column=1).value == self.id:
                folha.cell(row=li, column=4).value = self.nif
                folha.cell(row=li, column=5).value = self.nome
                folha.cell(row=li, column=6).value = self.morada
        livro.save(self.db)

    def passes(self):
        livro = openpyxl.load_workbook(self.db)
        folha = livro.active
        passes = {}
        for li in range(1, folha.max_row + 1):
            passes[folha.cell(row=li, column=1).value] = folha.cell(row=li, column=3).value
        return passes

    @staticmethod
    def code(passe):
        import hashlib
        return hashlib.sha3_256(passe.encode()).hexdigest()
